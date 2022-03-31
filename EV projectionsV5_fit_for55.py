#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Dec  4 11:20:56 2020

@author: ryanoreilly
The objective of this python script is to project the number of Electric vehicles into the future. I will use time series data on new vehicles by country from
https://www.acea.be/statistics/tag/category/by-country-registrations. 
Turkey's number of new vehicles came from: TurkeyStat https://data.tuik.gov.tr/Search/Search?text=vehicles&dil=2
Cyprus' number of new vehicles came from monthly reports by the ACEA. 
Additionally the share of new vehicle fleets that will be used will come from: https://www.transportenvironment.org/publications/hitting-ev-inflection-point  
Number of existing EVs come from EAFO: https://www.eafo.eu/countries/european-union/23640/summary

Difference between V4 and V5 is that the average between #years assuming 150k life and average age of fleet is used instead of # years assuming 150k life
also dropped IS from remove EV which was causing decomissioning of EV error

This file creates a scenario where all vehicle sales are electric starting in 2035
"""

import pandas as pd ## necessary data analysis package
import numpy as np ## necessary data analysis package
import os
import openpyxl as xl
### Set options
#os.chdir('I:\Projekte\OpenEntrance - WV0173\Durchführungsphase\WP6\CS1\gitlab\datainputs') # set wd - change to server path - IMPORTANT
os.chdir('I:/Projekte/OpenEntrance - WV0173/Durchführungsphase/WP6/CS1/OE_data_analysis/openEntrance/data')

## TEST TEST COMMMENT
# Display 6 columns for viewing purposes
pd.set_option('display.max_columns', 6)
# Reduce decimal points to 2
pd.options.display.float_format = '{:,.2f}'.format

#########################################
#import history of New passenger car registrations
#########################################
wb1 = xl.load_workbook("I:/Projekte/OpenEntrance - WV0173/Durchführungsphase/WP6/CS1/OE_data_analysis/openEntrance/data/openENTRANCE projection/ACEA - New Vehicle Registration/1990-2020_PC_by_country_EU+EFTA+UK.xlsx")
ws1 = wb1.worksheets[0]
TR_CY = pd.read_excel("I:/Projekte/OpenEntrance - WV0173/Durchführungsphase/WP6/CS1/OE_data_analysis/openEntrance/data/openENTRANCE projection/turkey_cyprus.xlsx")

#########################################
#import existing EV 
#########################################

EV08_20 = pd.read_excel("I:/Projekte/OpenEntrance - WV0173/Durchführungsphase/WP6/CS1/OE_data_analysis/openEntrance/data/openENTRANCE projection/EAFO Electric Vehicles 08-20/EAFO EV 08-20.xlsx", sheet_name="total BEV fleet")
EV08_20 = EV08_20.sort_values(by="country")
EV18_20 = EV08_20[['country',2018,2019,2020]]
del EV08_20

#########################################
#import NVF EV 08 20 
#########################################

NVF_EV08_20 = pd.read_excel("I:/Projekte/OpenEntrance - WV0173/Durchführungsphase/WP6/CS1/OE_data_analysis/openEntrance/data/openENTRANCE projection/EAFO Electric Vehicles 08-20/EAFO EV 08-20.xlsx",sheet_name = 'NVF BEV by year')
NVF_EV08_20.index = NVF_EV08_20['country']
NVF_EV08_20 = NVF_EV08_20.drop('country', axis = 1)
# drop redundent years
NVF_EV08_17 = NVF_EV08_20.drop([2018,2019,2020],axis = 1)
del NVF_EV08_20
#########################################
#usable life of EV-assuming 150,000km - create matrix of decomission of vehicles
#########################################
EVlife = pd.read_excel('I:/Projekte/OpenEntrance - WV0173/Durchführungsphase/WP6/CS1/OE_data_analysis/openEntrance/data/EV_parameters.xlsx')
#EVlife['#years150km'] = round(EVlife['evLIFE_150kkm'])
EVlife['#years150km'] = round(EVlife['average age/# years assuming 150k life']) # new method to account for low age of vehicles
EVlife.index = EVlife['country']
EVlife = EVlife.drop(['evLIFE_150kkm','wunit_EV','country','average age/# years assuming 150k life'], axis = 1)

columns = []
for i in range(2008,2051):
    columns.append(i)
years = list(range(2008,2051))
removeEV = pd.DataFrame(index = EVlife.index, columns = columns)

for nat in removeEV.index:
    life = EVlife.loc[nat]
    for year in years:
        removeEV.loc[nat][year]     = int(year + life)
#########################################
#import expected path of EV registration rates of New passenger car registrations
#########################################     
#from T&E 2021 = shares up to 2035 - all shares are assumed to be 100% (all NVF are EV) to draw a scenario for the fit for 55
rEV_path =  pd.read_excel("I:/Projekte/OpenEntrance - WV0173/Durchführungsphase/WP6/CS1/OE_data_analysis/openEntrance/data/EV_NVF_EV_path.xlsx", sheet_name = "fit_for_55")
#rEV_path =  pd.read_excel("C:/Users/AK194059/Desktop/Ev lit review.xlsx", sheet_name = "2021 T&E final")
indexCountries = rEV_path[['NAME','country']] # used to index the full name to the acronyms below\
indexCountries['name'] = indexCountries['NAME'].str.lower() # make names all lowercase
rEV_path = rEV_path.drop('NAME', axis = 1)


########################################
#######################################
# NEW VEHICLE FLEET PROJECTIONS
#######################################
#######################################

#########################################
#extract annual new vehicle registrations for each country
#########################################
country = []
newV = []
year = []    

yr = 1989
for j in range(0,31):
    yr +=1
    print("yr: ",yr)
    #call worksheet
    ws1 = wb1.worksheets[j]
    # columns in source excel file 
    mr = ws1.max_row 
    for i in range(1,mr+1):
        if  ws1.cell(row = i,column=1).value == "Austria":
            Frow = i
            print("Frow: ",i)
        if  ws1.cell(row = i,column=1).value == "SOURCE: ASSOCIATION AUXILIAIRE DE L'AUTOMOBILE (AAA)":
            Lrow = i
            print("Lrow: ",i)
    for a in range(Frow,Lrow):
        temp = (ws1.cell(row = a, column=1).value)
        country.append(temp.lower())
        newV.append(ws1.cell(row = a, column=14).value)
        year.append(yr)
 
    
vehicleTS = pd.DataFrame()
vehicleTS['country'] = country
vehicleTS['newV'] = newV
vehicleTS['year'] = year
# concat Turkey and Cyprus
vehicleTS = pd.concat([vehicleTS, TR_CY])
vehicleTS['growth'] = ""

#remove romania and change romania1 to romania
index_names = vehicleTS[vehicleTS['country'] == "romania"].index
vehicleTS.drop(index_names, inplace=True)
vehicleTS['country'] =vehicleTS['country'].replace(['romania1'],"romania")

# since 2020 was not a normal year, we will drop it for the calculation of the average annual NVF
temp = vehicleTS[vehicleTS['year'] != 2020]
temp2 = temp[temp['year'] >= 2013]

summary=temp2.groupby(['country']).mean()
summary = summary.sort_values(by="country")
#########################################
#project new vehicles registrations through 2050 using mean new vehicle registrations
#########################################
columns = []
for i in range(2021,2051):
    columns.append(i)#str(i))
dfP = pd.DataFrame(index = summary.index, columns = columns)   
    

for row in range(0,len(dfP.index)):
    for col in range(0,dfP.shape[1]):
        dfP.iloc[row,col] = summary['newV'].iloc[row] 


########################################
#######################################
# EV VEHICLE FLEET PROJECTIONS
#######################################
#######################################

#########################################
# fill in the future expected EV NVF
#########################################   

# create column of the countries of interest
temp = pd.merge(indexCountries,summary, left_on="name", right_on=summary.index)

#extract data for new vehicle registrations 2018, 2019, 2020
vehicleTS=vehicleTS[['country','year','newV']]
NVF = vehicleTS.pivot(index='year', columns='country', values='newV')
NVF.iloc[0].isna
NVF =  pd.DataFrame(NVF).T
NVF = NVF[[2018,2019,2020]]
# merge existing data with projected data of new vehicle registrations
NVF= pd.merge(temp, NVF,left_on="name", right_on="country")
NVF = NVF[['name','country',2018,2019,2020]]
NVF = pd.merge(NVF, dfP, left_on='name', right_on='country')
# NVF for TR and CY are missing for 2020 -> we will use the average

NVF.index = NVF['country']
NVF= NVF.drop(['country','name'], axis=1)

columns = []
for i in range(2018,2051):
    columns.append(i)
yrEV = pd.DataFrame(index = rEV_path['country'], columns = columns)
# set column 'country' as index
rEV_path.index = rEV_path['country']
rEV_path= rEV_path.drop(['country'], axis=1)   

#fill in yrEV
for row in yrEV.index:
    temp2 = []
    for col in range(2018,2051):
        temp2.append(round(rEV_path.loc[row][col]*float(NVF.loc[row][col]),0))
    yrEV.loc[row] = temp2

#save to file
yrEV.to_csv(r'EV NVF projectionsV2_fit55.csv')  
NVF.to_csv(r'New Vehicle Fleet projectionsV2.csv')  
#df_rEV.to_csv(r'rates of EV of NVF.csv')
#save existing time series to file
vehicleTS=vehicleTS[['country','year','newV']]

vehicleTS = vehicleTS.pivot(index='year', columns='country', values='newV')
vehicleTS.iloc[0].isna
vehicleTS =  pd.DataFrame(vehicleTS).T
vehicleTS.to_csv(r'./openENTRANCE projection/ACEA - New Vehicle Registration/Historical Vehicle Registration 1990-2020V2.csv') 


#########################################
#import national household projections
#########################################
nhh = pd.read_csv('./EUROSTAT POP PROJ/NATnhhV2.csv', sep=",",encoding = "ISO-8859-1", header=0, index_col=False)
nhh.index = nhh['country']
nhh = nhh.drop(['country','Unnamed: 0'], axis = 1)
nhh = nhh.drop(labels='MT', axis = 0 ) #drop Malta

############################################
#formula region - n; year - i; 

# #EV(n,i) = #EV(i-1) + NVF_EV(n,i) + (NVF(n,2018)/NHH(n,2018) * (NHH(n,i)-NHH(n,i-1)*r_EV)) - #EV_decomission(n,i)

# #EV      = #EV previous year + new EVs added based on market expectations + # EVs added based on change in # households - #EVs that are decomissioned in year i

##########################################
# NVF_EV 2008-2050
NVF_EV08_50 = pd.merge(NVF_EV08_17, yrEV, left_on= NVF_EV08_17.index, right_on=yrEV.index) 
NVF_EV08_50.index = NVF_EV08_50['key_0']
NVF_EV08_50 = NVF_EV08_50.drop('key_0',axis=1)

columns = []
for i in range(2021,2051):
    columns.append(i)
EV = pd.DataFrame(index = temp['country'], columns = columns)   # NVF_EV added through expectations in markets and through change in #HH less decomissioned EVs
EV_NHH = pd.DataFrame(index = temp['country'], columns = columns)   # NVF_EV added through change in #HH

EV = pd.merge(EV18_20, EV,left_on = "country", right_on = "country")
EV.index = EV.country
EV = EV.drop('country',axis=1)
EV = EV.sort_index()
EV_NHH = EV_NHH.sort_index()
yrEV = yrEV.sort_index() #NVF EV
#remove iceland
EV = EV.drop('IS', axis =0)
EV_NHH =  EV_NHH.drop('IS', axis =0)
yrEV = yrEV.drop('IS', axis =0)
NVF_EV08_50 =  NVF_EV08_50.drop('IS', axis =0)
NVF = NVF.drop('IS', axis = 0)
rEV_path = rEV_path.drop('IS', axis = 0)
removeEV = removeEV.drop('IS', axis = 0)

EV_rem = pd.DataFrame(index = EV.index, columns = columns)   # annual decommissioned EVs
EV_rem = EV_rem.sort_index()

#check indexes for consistency
#set(rEV_path.index) == set(EV.index)


for row in range(0,len(EV.index)):
    print("country",EV.index[row])
    for col in range(3,EV.shape[1]):
        print("year",list(EV.columns)[col])
        if  (NVF.iloc[row,1]/nhh.iloc[row,1]*(nhh.iloc[row,col]-nhh.iloc[row,col-1])*rEV_path.iloc[row,col]) <0: # if the EV NVF that is added through change in HH is negative do not add to the respective cell in NVF_EV08_50, as these vehicles have allready been removed
            print("NHH_noREM")
        else: # if the EV NVF that is added through change in HH is positive, add value to the respective cell in NVF_EV08_50, as these vehicles will have to be decomissioned in future years
            print("NHH_REM!!!!","country: ",EV.index[row], "year: ",list(EV.columns)[col],"NVF current: ",NVF_EV08_50.loc[EV.index[row],list(EV.columns)[col]]," + ",(NVF.iloc[row,1]/nhh.iloc[row,1]*(nhh.iloc[row,col]-nhh.iloc[row,col-1])*rEV_path.iloc[row,col]))                    
            NVF_EV08_50.loc[EV.index[row],list(EV.columns)[col]] = NVF_EV08_50.loc[EV.index[row],list(EV.columns)[col]] + (NVF.iloc[row,1]/nhh.iloc[row,1]*(nhh.iloc[row,col]-nhh.iloc[row,col-1])*rEV_path.iloc[row,col])        
        for rem in range(0,44):
        #while rem !=  43:#removeEV.iloc[row,rem]:
            #rem +=1
            #print("rem",rem)
            if rem == 43: # if activated no removal of EV in year i
                print("IF: ", rem)
                EV.iloc[row,col]       = int(EV.iloc[row,col-1]+yrEV.iloc[row,col] + (NVF.iloc[row,1]/nhh.iloc[row,1]*(nhh.iloc[row,col]-nhh.iloc[row,col-1])*rEV_path.iloc[row,col]))
                EV_NHH.iloc[row,col-3] = ((NVF.iloc[row,1]/nhh.iloc[row,1])*(nhh.iloc[row,col]-nhh.iloc[row,col-1])*rEV_path.iloc[row,col]) 
                print("#EV i -1",EV.iloc[row,col-1],"*")
                print("#NVF EV",yrEV.iloc[row,col] )
                print("#remove",0)
                print("#EV",EV.iloc[row,col],"~~~")
                break 
            elif col+2018 ==  removeEV.iloc[row,rem]: # if activated removal of EV in year i
                print("ELIF: ", rem)
                remove = NVF_EV08_50.iloc[row,rem]
                EV_rem.iloc[row,col-3] = remove
                EV.iloc[row,col]       = int(EV.iloc[row,col-1]+yrEV.iloc[row,col] + (NVF.iloc[row,1]/nhh.iloc[row,1]*(nhh.iloc[row,col]-nhh.iloc[row,col-1])*rEV_path.iloc[row,col])-remove)
                EV_NHH.iloc[row,col-3] = (NVF.iloc[row,1]/nhh.iloc[row,1]*(nhh.iloc[row,col]-nhh.iloc[row,col-1])*rEV_path.iloc[row,col])
                print("#EV i -1",EV.iloc[row,col-1],"*")
                print("#NVF EV",yrEV.iloc[row,col] )
                print("#remove",remove)
                print("#EV",EV.iloc[row,col],"~~~")
                break        


#save to file
EV_NHH.to_csv(r'EV from NHHV5_ave_fit55.csv')  
EV_rem.to_csv(r'EV discontinuedV5_ave_fit55.csv')  
EV.to_csv(r'EV projectionsV5_ave_fit55.csv')  


